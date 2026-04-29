"""
Geocode sanitized addresses using Nominatim API.

This script takes the sanitized addresses (column S) and retrieves coordinates
from Nominatim's free geocoding service. Results are stored in columns T (latitude)
and U (longitude).

Important: Be respectful of Nominatim's usage policy:
- Add delays between requests (1 second minimum)
- Include User-Agent header
- Cache results to avoid duplicate requests
"""

import time
import json
import openpyxl
from pathlib import Path
from typing import Optional, Tuple, Dict
import requests
from urllib.parse import quote


class NominatimGeocoder:
    """Geocode addresses using Nominatim API."""
    
    BASE_URL = "https://nominatim.openstreetmap.org/search.php"
    
    def __init__(self, delay: float = 1.5, timeout: int = 10):
        """
        Initialize geocoder.
        
        Args:
            delay: Delay between requests in seconds (min 1.0 to respect usage policy)
            timeout: Request timeout in seconds
        """
        self.delay = max(delay, 1.0)
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'ruriosmon@gmail.com'
        })
        self.cache: Dict[str, Tuple[Optional[float], Optional[float]]] = {}
        self.last_request_time = 0
    
    def geocode(self, address: str) -> Tuple[Optional[float], Optional[float]]:
        """
        Geocode a single address to (latitude, longitude).
        
        Args:
            address: Address string to geocode
        
        Returns:
            Tuple of (latitude, longitude) or (None, None) if not found
        """
        if not address or not isinstance(address, str):
            return None, None
        
        # Check cache first
        if address in self.cache:
            return self.cache[address]
        
        # Respect rate limiting
        elapsed = time.time() - self.last_request_time
        if elapsed < self.delay:
            time.sleep(self.delay - elapsed)
        
        try:
            params = {
                'q': address,
                'format': 'json',
                'limit': 1  # Get best match only
            }
            
            response = self.session.get(self.BASE_URL, params=params, timeout=self.timeout)
            response.raise_for_status()
            
            self.last_request_time = time.time()
            
            results = response.json()
            if results:
                lat = float(results[0].get('lat'))
                lon = float(results[0].get('lon'))
                self.cache[address] = (lat, lon)
                return lat, lon
            else:
                self.cache[address] = (None, None)
                return None, None
        
        except requests.exceptions.RequestException as e:
            print(f"Error geocoding '{address}': {e}")
            return None, None
    
    def process_xlsx(self, xlsx_file: Path, 
                    input_column: int = 19,      # Column S: sanitized addresses
                    lat_column: int = 20,        # Column T: latitude
                    lon_column: int = 21,        # Column U: longitude
                    skip_existing: bool = True,
                    verbose: bool = True) -> Tuple[int, int]:
        """
        Process an Excel file: geocode addresses and save coordinates.
        
        Args:
            xlsx_file: Path to Excel file
            input_column: Column with sanitized addresses (1-indexed)
            lat_column: Column for latitude results (1-indexed)
            lon_column: Column for longitude results (1-indexed)
            skip_existing: Skip cells that already have coordinates
            verbose: Print progress
        
        Returns:
            Tuple of (geocoded_count, total_count)
        """
        wb = openpyxl.load_workbook(xlsx_file)
        
        total = 0
        geocoded = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Iterate with max_col to ensure we get all columns we need
            for row in ws.iter_rows(min_row=2, max_col=max(input_column, lat_column, lon_column)):
                address_cell = row[input_column - 1]
                lat_cell = row[lat_column - 1]
                lon_cell = row[lon_column - 1]
                
                # Skip if no address
                if not address_cell.value:
                    continue
                
                total += 1
                
                # Skip if coordinates already exist
                if skip_existing and lat_cell.value and lon_cell.value:
                    continue
                
                address = str(address_cell.value)
                lat, lon = self.geocode(address)
                
                if lat is not None and lon is not None:
                    ws.cell(row=row[0].row, column=lat_column, value=round(lat, 6))
                    ws.cell(row=row[0].row, column=lon_column, value=round(lon, 6))
                    geocoded += 1
                    
                    if verbose and geocoded % 50 == 0:
                        print(f"  Progress: {geocoded}/{total} addresses geocoded")
                
        wb.save(xlsx_file)
        return geocoded, total


def main():
    """Geocode all result files using Nominatim."""
    
    results_dir = Path("Resultats")
    geocoder = NominatimGeocoder(delay=1.5)
    
    xlsx_files = sorted(results_dir.glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"No XLSX files found in {results_dir}")
        return
    
    print("=" * 70)
    print("NOMINATIM GEOCODING - Processing Historical Lausanne Addresses")
    print("=" * 70)
    print(f"\nFound {len(xlsx_files)} files to process")
    print("Adding coordinates to columns T (latitude) and U (longitude)")
    print("This will take some time due to rate limiting (1.5s per request)\n")
    
    total_geocoded = 0
    total_addresses = 0
    
    for xlsx_file in xlsx_files:
        #if xlsx_file.name != "prelaz1951.xlsx":  # Skip already geocoded file
            #continue
        print(f"\nProcessing: {xlsx_file.name}")
        try:
            geocoded, total = geocoder.process_xlsx(xlsx_file)
            total_geocoded += geocoded
            total_addresses += total
            print(f"  ✓ Geocoded {geocoded}/{total} addresses")
        except Exception as e:
            print(f"  ✗ Error: {e}")
    
    print("\n" + "=" * 70)
    print(f"COMPLETE: Geocoded {total_geocoded}/{total_addresses} addresses")
    print("=" * 70)
    print("\nCoordinates stored in:")
    print("  - Column T: Latitude")
    print("  - Column U: Longitude")
    print("\nNext steps:")
    print("1. Create a GeoJSON file from the coordinates")
    print("2. Load into QGIS or a web mapping library")
    print("3. Visualize the evolution of neighborhoods")


if __name__ == "__main__":
    main()
