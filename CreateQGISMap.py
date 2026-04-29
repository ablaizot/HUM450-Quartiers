"""
Create GeoJSON files from geocoded addresses for QGIS visualization.

This script exports all geocoded addresses to GeoJSON format, ready to be
imported into QGIS for mapping and analysis.
"""

import json
import openpyxl
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import re


class GeoJSONExporter:
    """Export geocoded data to GeoJSON format."""
    
    def __init__(self):
        self.features: List[Dict[str, Any]] = []
    
    def add_point(self, lat: float, lon: float, properties: Dict[str, Any]):
        """Add a point feature."""
        if lat is None or lon is None:
            return
        
        feature = {
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [lon, lat]
            },
            "properties": properties
        }
        self.features.append(feature)
    
    def to_geojson(self) -> Dict[str, Any]:
        """Return the complete GeoJSON FeatureCollection."""
        return {
            "type": "FeatureCollection",
            "features": self.features
        }
    
    def save(self, output_file: Path):
        """Save GeoJSON to file."""
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.to_geojson(), f, indent=2, ensure_ascii=False)
        print(f"    Saved: {output_file}")


def process_excel_file(xlsx_file: Path, output_dir: Path) -> tuple:
    """
    Extract geocoded data from Excel and create GeoJSON.
    
    Returns: (geojson_path, num_features, year)
    """
    exporter = GeoJSONExporter()
    
    wb = openpyxl.load_workbook(xlsx_file)
    count = 0
    year = None
    
    # Extract year from filename
    match = re.search(r'(\d{4})', xlsx_file.name)
    if match:
        year = int(match.group(1))
    
    # Process all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=2, max_col=21):
            # Columns: S=address(19), T=lat(20), U=lon(21)
            address_cell = row[18]  # Column S
            lat_cell = row[19]      # Column T
            lon_cell = row[20]      # Column U
            
            lat = lat_cell.value
            lon = lon_cell.value
            address = address_cell.value
            
            if lat and lon and address:
                try:
                    lat_float = float(lat)
                    lon_float = float(lon)
                    
                    props = {
                        "address": str(address),
                        "sheet": sheet_name,
                    }
                    if year:
                        props["year"] = year
                    
                    exporter.add_point(lat_float, lon_float, props)
                    count += 1
                except (ValueError, TypeError):
                    pass
    
    # Save GeoJSON
    stem = xlsx_file.stem
    geojson_file = output_dir / f"{stem}.geojson"
    exporter.save(geojson_file)
    
    return geojson_file, count, year


def main():
    """Create GeoJSON files from geocoded data."""
    results_dir = Path("Resultats")
    output_dir = Path("GeoJSON_Output")
    output_dir.mkdir(exist_ok=True)
    
    xlsx_files = sorted(results_dir.glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"No XLSX files found in {results_dir}")
        return
    
    print("=" * 80)
    print("EXPORTING TO GEOJSON - For QGIS Visualization")
    print("=" * 80)
    print(f"\nFound {len(xlsx_files)} files to process\n")
    
    total_features = 0
    
    # Process each file
    for xlsx_file in xlsx_files:
        print(f"Processing: {xlsx_file.name}")
        try:
            geojson_file, count, year = process_excel_file(xlsx_file, output_dir)
            total_features += count
            print(f"  ✓ Exported {count} points (year: {year})")
            
        except Exception as e:
            print(f"  ✗ Error: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n{'='*80}")
    print(f"COMPLETE: Exported {total_features} points total")
    print(f"{'='*80}\n")
    
    print(f"GeoJSON files created in: {output_dir}/\n")
    
    print("TO VISUALIZE IN QGIS:")
    print("-" * 80)
    print("1. Open QGIS Desktop")
    print("2. Go to: Layer → Add Layer → Add Vector Layer")
    print("3. Source: File")
    print("4. Browse to: GeoJSON_Output/")
    print("5. Select all .geojson files (Ctrl+A)")
    print("6. Click 'Add'")
    print("\nTO STYLE BY YEAR:")
    print("-" * 80)
    print("7. Right-click a layer in the Layers panel")
    print("8. Click 'Properties'")
    print("9. Go to 'Symbology' tab")
    print("10. Change 'Single Symbol' to 'Categorized'")
    print("11. Column: 'year'")
    print("12. Click 'Classify'")
    print("13. Double-click each color to customize")
    print("\nCOLOR SCHEME BY YEAR (for temporal visualization):")
    print("-" * 80)
    print("  1873-1890: Blue (#0000FF) → Cyan (#00CCFF)     [Early period]")
    print("  1900-1912: Green (#55FF00) → Yellow (#FFFF00)  [Middle period]")
    print("  1923-1951: Orange (#FF9900) → Dark Red (#990000) [Recent period]")
    print("\nTIPS:")
    print("-" * 80)
    print("• Use 'Identify' tool (click on points) to see address details")
    print("• Adjust transparency to see overlapping points")
    print("• Use 'Select by Expression' to filter by year or neighborhood")
    print("• Export as PDF for reports")


if __name__ == "__main__":
    main()
