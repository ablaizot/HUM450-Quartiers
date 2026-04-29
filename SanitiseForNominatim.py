"""
Sanitize addresses from historical directory data for Nominatim geocoding.

This script processes addresses extracted from historical directories (1880s-1950s)
and prepares them for geocoding via Nominatim.

Strategy:
1. Use the sheet name as the primary street/location (manually verified)
2. Extract house number from OCR text if present
3. Combine: street name + number (if any) + city context

This approach is robust to OCR errors since we rely on the reliable sheet name
rather than trying to parse OCR-corrupted address text.
"""

import re
import openpyxl
from pathlib import Path
from typing import List, Optional, Tuple


class AddressSanitizer:
    """Sanitize historical addresses for Nominatim geocoding."""
    
    def __init__(self, city: str = "Lausanne", country: str = "Switzerland"):
        """
        Initialize the sanitizer.
        
        Args:
            city: City name (default: Lausanne)
            country: Country name (default: Switzerland)
        """
        self.city = city
        self.country = country
    
    def extract_house_number(self, text: str) -> Optional[str]:
        """
        Extract house/street number from OCR text.
        
        Looks for patterns like:
        - "123" at the start
        - "123bis" or "123ter"
        - "n°123" or "no123"
        - "rue de X 123" (number at end)
        
        Args:
            text: OCR-extracted address text
        
        Returns:
            House number as string, or None if not found
        """
        if not text:
            return None
        
        # Remove leading articles and clean up
        text = str(text).strip().lower()
        
        # Remove list brackets and quotes
        text = text.strip('[]\'\"')
        
        # Pattern 1: n°123, no123, n.123, etc.
        match = re.search(r'n[°o]\.?\s*(\d+(?:bis|ter|quater)?)', text)
        if match:
            return match.group(1).capitalize()
        
        # Pattern 2: number at the very start (sometimes OCR sticks it to text)
        match = re.match(r'^(\d+(?:bis|ter|quater)?)', text)
        if match:
            return match.group(1).capitalize()
        
        # Pattern 3: number with separators before text (e.g., "123-rue", "123 rue")
        match = re.match(r'^(\d+(?:bis|ter|quater)?)\s*[-,/]?\s*[a-z]', text)
        if match:
            return match.group(1).capitalize()
        
        # Pattern 4: number at the end (e.g., "rue de la paix 123")
        # This is the most common case in our OCR text
        match = re.search(r'(?:^|\s)(\d+(?:bis|ter|quater)?)\s*$', text)
        if match:
            return match.group(1).capitalize()
        
        # Pattern 5: number followed by space then end (alternative)
        match = re.search(r'\s(\d+(?:bis|ter|quater)?)\s*$', text)
        if match:
            return match.group(1).capitalize()
        
        return None
    
    def sanitize_street_name(self, sheet_name: str) -> str:
        """
        Clean and capitalize a street name from sheet name.
        
        Args:
            sheet_name: Name from Excel sheet
        
        Returns:
            Cleaned street name
        """
        if not sheet_name:
            return ""
        
        # Normalize whitespace
        name = re.sub(r'\s+', ' ', sheet_name.strip())
        
        # Capitalize properly: first word and proper nouns
        words = name.split()
        capitalized = []
        
        for i, word in enumerate(words):
            word_lower = word.lower()
            
            # First word: always capitalize
            if i == 0:
                capitalized.append(word.capitalize())
            # Small prepositions: keep lowercase
            elif word_lower in ('de', 'du', 'des', 'la', 'le', 'les', 'et', 'ou', 'à'):
                capitalized.append(word_lower)
            # Numbers: keep as-is
            elif word[0].isdigit():
                capitalized.append(word)
            # Everything else: capitalize
            else:
                capitalized.append(word.capitalize())
        
        return ' '.join(capitalized)
    
    def combine_address(self, street_name: str, house_number: Optional[str] = None) -> str:
        """
        Combine street name and house number into final address.
        
        Args:
            street_name: Cleaned street name
            house_number: House number or None
        
        Returns:
            Complete address for Nominatim
        """
        if house_number:
            return f"{house_number} {street_name}, {self.city}, {self.country}"
        else:
            return f"{street_name}, {self.city}, {self.country}"
    
    def sanitize_from_sheet(self, ocr_text: str, sheet_name: str) -> str:
        """
        Sanitize address using sheet name and OCR text.
        
        Args:
            ocr_text: OCR-extracted text from cell
            sheet_name: Street name from Excel sheet
        
        Returns:
            Cleaned address ready for Nominatim
        """
        if not sheet_name:
            return ""
        
        # Extract house number from OCR text
        house_number = self.extract_house_number(ocr_text)
        
        # Clean the street name
        street_name = self.sanitize_street_name(sheet_name)
        
        # Combine into final address
        return self.combine_address(street_name, house_number)
    
    def process_xlsx(self, xlsx_file: Path, output_file: Optional[Path] = None,
                     address_column: int = 4, output_column: int = 19) -> int:
        """
        Process an Excel file with addresses using sheet names.
        
        Args:
            xlsx_file: Path to input XLSX file
            output_file: Path to output XLSX file (if None, overwrites input)
            address_column: Column index containing OCR text (1-indexed)
            output_column: Column index for sanitized addresses (1-indexed)
        
        Returns:
            Number of addresses processed
        """
        wb = openpyxl.load_workbook(xlsx_file)
        
        count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2):  # Skip header
                cell = row[address_column - 1]
                if cell.value:
                    # Use sheet name as street, extract number from cell
                    sanitized = self.sanitize_from_sheet(str(cell.value), sheet_name)
                    ws.cell(row=row[0].row, column=output_column, value=sanitized)
                    count += 1
        
        if output_file is None:
            output_file = xlsx_file
        
        wb.save(output_file)
        return count


def main():
    """Process all result files to add sanitized addresses for Nominatim."""
    
    results_dir = Path("Resultats")
    sanitizer = AddressSanitizer()
    
    xlsx_files = sorted(results_dir.glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"No XLSX files found in {results_dir}")
        return
    
    print(f"Found {len(xlsx_files)} files to process")
    print("-" * 70)
    
    for xlsx_file in xlsx_files:
        print(f"\nProcessing: {xlsx_file.name}")
        try:
            count = sanitizer.process_xlsx(xlsx_file, address_column=4, output_column=19)
            print(f"  ✓ Sanitized {count} addresses")
            print(f"  Using sheet names as street names + extracted house numbers")
            print(f"  Added to column S (19)")
        except Exception as e:
            print(f"  ✗ Error: {e}")
    
    print("\n" + "-" * 70)
    print("Done! Sanitized addresses are in column S of each file.")
    print("\nApproach:")
    print("  • Sheet name = Official street name (manually verified)")
    print("  • Extracted number from OCR text = House number (if present)")
    print("  • Result = 'Street Name' or 'Number Street Name', Lausanne, Switzerland")
    print("\nNext steps:")
    print("1. Use Nominatim API to geocode addresses in column S")
    print("2. Example: https://nominatim.openstreetmap.org/search.php?q={address}&format=json")
    print("3. Save coordinates to columns T (latitude) and U (longitude)")


if __name__ == "__main__":
    main()


def main():
    """Process all result files to add sanitized addresses for Nominatim."""
    
    results_dir = Path("Resultats")
    results_dest_dir = Path("ResultatsVraiAddr")
    sanitizer = AddressSanitizer()
    
    xlsx_files = sorted(results_dir.glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"No XLSX files found in {results_dir}")
        return
    
    print(f"Found {len(xlsx_files)} files to process")
    print("-" * 60)
    
    for xlsx_file in xlsx_files:
        print(f"\nProcessing: {xlsx_file.name}")
        try:
            count = sanitizer.process_xlsx(xlsx_file, address_column=4, output_column=19)
            print(f"  ✓ Sanitized {count} addresses")
            print(f"  Added to column S (19)")
        except Exception as e:
            print(f"  ✗ Error: {e}")
    
    print("\n" + "-" * 60)
    print("Done! Sanitized addresses are in column S of each file.")
    print("\nNext steps:")
    print("1. Use Nominatim API to geocode addresses in column S")
    print("2. Example: https://nominatim.openstreetmap.org/search.php?q={address}&format=json")
    print("3. Save coordinates to columns T (latitude) and U (longitude)")


if __name__ == "__main__":
    main()
